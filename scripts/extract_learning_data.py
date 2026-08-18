"""Convert course workbooks into browser-friendly prototype data."""

from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "src" / "data"

COURSES = [
    {
        "id": "math",
        "source": ROOT / "ФИНАЛ для прототипа ии_тьютора.xlsx",
        "output": DATA_DIR / "learning-data-math.json",
        "format": "math",
    },
    {
        "id": "fingram",
        "source": ROOT / "AI Tutor контент - Финграм.xlsx",
        "output": DATA_DIR / "learning-data-fingram.json",
        "format": "fingram",
    },
]

CORRECT_MARKERS = {
    "1",
    "true",
    "да",
    "верно",
    "+",
    "yes",
}


def normalized(value: object) -> str:
    text = str(value or "").translate(
        {
            ord("\u200b"): None,
            ord("\u200c"): None,
            ord("\u200d"): None,
            ord("\ufeff"): None,
        }
    )
    return " ".join(text.split())


def clean_content(value: object) -> str:
    return str(value or "").translate(
        {
            ord("\u200b"): None,
            ord("\u200c"): None,
            ord("\u200d"): None,
            ord("\ufeff"): None,
        }
    ).strip()


def rows(workbook, sheet: str):
    worksheet = workbook[sheet]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if any(value is not None for value in row):
            yield row


def description_from_theory(theory: str) -> str:
    for line in theory.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        return stripped
    return ""


def is_correct_marker(value: object) -> bool:
    return str(value or "").strip().lower() in CORRECT_MARKERS


def build_payload(
    *,
    course_id: str,
    source: Path,
    skill_names: list[str],
    descriptions: dict[str, str],
    theories: dict[str, str],
    prerequisites: dict[str, list[str]],
    next_skills: dict[str, list[str]],
    task_skills: dict[str, list[str]],
    open_answers: dict[str, str],
    test_answers: dict[str, list[dict[str, object]]],
) -> dict[str, object]:
    all_names = list(
        dict.fromkeys(
            skill_names
            + list(descriptions)
            + list(theories)
            + list(prerequisites)
            + [name for values in prerequisites.values() for name in values]
            + [name for values in task_skills.values() for name in values]
        )
    )
    ids = {name: f"s{index:04d}" for index, name in enumerate(all_names, 1)}

    tasks_by_skill: dict[str, list[dict[str, object]]] = defaultdict(list)
    for index, (text, skills) in enumerate(task_skills.items(), 1):
        options = test_answers.get(text, [])
        answer = open_answers.get(text, "")
        if not answer and options:
            answer = next(
                (option["text"] for option in options if option.get("correct")),
                "",
            )
        task = {
            "id": f"t{index:04d}",
            "text": text,
            "answer": answer,
            "options": options,
        }
        for skill in skills:
            tasks_by_skill[skill].append(task)

    skills = []
    for name in all_names:
        theory = theories.get(name, "")
        tasks = tasks_by_skill.get(name, [])
        description = descriptions.get(name, "") or description_from_theory(theory)
        skills.append(
            {
                "id": ids[name],
                "title": name,
                "description": description,
                "theory": theory,
                "prerequisites": [
                    ids[item]
                    for item in dict.fromkeys(prerequisites.get(name, []))
                    if item in ids
                ],
                "next": [
                    ids[item]
                    for item in dict.fromkeys(next_skills.get(name, []))
                    if item in ids
                ],
                "tasks": tasks,
                "search": normalized(f"{name} {description}").lower(),
            }
        )

    featured = next(
        (
            ids[name]
            for name in all_names
            if theories.get(name) and len(tasks_by_skill.get(name, [])) >= 3
        ),
        skills[0]["id"] if skills else "",
    )

    return {
        "meta": {
            "courseId": course_id,
            "source": source.name,
            "skills": len(skills),
            "theories": sum(bool(skill["theory"]) for skill in skills),
            "tasks": len(task_skills),
            "relations": sum(len(skill["prerequisites"]) for skill in skills),
            "fingerprint": hashlib.sha1(source.read_bytes()).hexdigest()[:10],
        },
        "featuredSkillId": featured,
        "skills": skills,
    }


def extract_math(workbook) -> dict[str, object]:
    skill_names = [normalized(row[0]) for row in rows(workbook, "Список навыков")]
    skill_names = [name for name in skill_names if name]
    descriptions = {
        normalized(row[0]): normalized(row[1])
        for row in rows(workbook, "Навык+описание")
        if row[0]
    }
    theories = {
        normalized(row[0]): clean_content(row[1])
        for row in rows(workbook, "Теория-Навык")
        if row[0] and row[1]
    }

    prerequisites: dict[str, list[str]] = defaultdict(list)
    next_skills: dict[str, list[str]] = defaultdict(list)
    for row in rows(workbook, "Пререквизитные отношения"):
        skill, prerequisite = normalized(row[0]), normalized(row[1])
        if skill and prerequisite:
            prerequisites[skill].append(prerequisite)
            next_skills[prerequisite].append(skill)

    open_answers = {
        normalized(row[0]): normalized(row[1])
        for row in rows(workbook, "Ответы открытые задания")
        if row[0] and row[1]
    }
    test_answers: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows(workbook, "Ответы тест"):
        task, option, marker = normalized(row[0]), normalized(row[1]), row[2]
        if task and option:
            test_answers[task].append(
                {"text": option, "correct": is_correct_marker(marker)}
            )

    task_skills: dict[str, list[str]] = defaultdict(list)
    for row in rows(workbook, "Задача - Навык"):
        task, skill = normalized(row[0]), normalized(row[1])
        if task and skill and skill not in task_skills[task]:
            task_skills[task].append(skill)

    return build_payload(
        course_id="math",
        source=COURSES[0]["source"],
        skill_names=skill_names,
        descriptions=descriptions,
        theories=theories,
        prerequisites=prerequisites,
        next_skills=next_skills,
        task_skills=task_skills,
        open_answers=open_answers,
        test_answers=test_answers,
    )


def extract_fingram(workbook) -> dict[str, object]:
    skill_rows = list(rows(workbook, "Список навыков"))
    skill_names = [
        normalized(row[0])
        for row in sorted(
            skill_rows,
            key=lambda item: float(item[1] or 0) if item[1] is not None else 0,
        )
        if row[0]
    ]

    theories = {
        normalized(row[0]): clean_content(row[2])
        for row in rows(workbook, "Навыки+ Теория")
        if row[0] and row[2]
    }
    descriptions = {
        name: description_from_theory(theory)
        for name, theory in theories.items()
    }

    prerequisites: dict[str, list[str]] = defaultdict(list)
    next_skills: dict[str, list[str]] = defaultdict(list)
    for row in rows(workbook, "Пререквизитные отношения"):
        skill, prerequisite = normalized(row[0]), normalized(row[2])
        if skill and prerequisite:
            prerequisites[skill].append(prerequisite)
            next_skills[prerequisite].append(skill)

    open_answers: dict[str, str] = {}
    task_skills: dict[str, list[str]] = defaultdict(list)
    for row in rows(workbook, "Открытые задания"):
        skill, text = normalized(row[0]), clean_content(row[2])
        if not skill or not text:
            continue
        answer = next(
            (normalized(value) for value in row[5:8] if value),
            "",
        )
        open_answers[text] = answer
        if skill not in task_skills[text]:
            task_skills[text].append(skill)

    test_answers: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows(workbook, "Задания с выбором"):
        skill, text, option, marker = (
            normalized(row[0]),
            clean_content(row[2]),
            normalized(row[5]),
            row[6],
        )
        if not skill or not text or not option:
            continue
        test_answers[text].append(
            {"text": option, "correct": is_correct_marker(marker)}
        )
        if skill not in task_skills[text]:
            task_skills[text].append(skill)

    return build_payload(
        course_id="fingram",
        source=COURSES[1]["source"],
        skill_names=skill_names,
        descriptions=descriptions,
        theories=theories,
        prerequisites=prerequisites,
        next_skills=next_skills,
        task_skills=task_skills,
        open_answers=open_answers,
        test_answers=test_answers,
    )


EXTRACTORS = {
    "math": extract_math,
    "fingram": extract_fingram,
}


def write_payload(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def main() -> None:
    for course in COURSES:
        workbook = load_workbook(course["source"], read_only=True, data_only=True)
        payload = EXTRACTORS[course["format"]](workbook)
        write_payload(course["output"], payload)
        print(
            f"Wrote {course['output'].relative_to(ROOT)}: "
            f"{payload['meta']['skills']} skills, "
            f"{payload['meta']['tasks']} tasks, "
            f"{payload['meta']['relations']} relations"
        )


if __name__ == "__main__":
    main()
